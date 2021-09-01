from django.shortcuts import render, redirect
from django.views import generic
from django.views.generic import CreateView, DeleteView
from django.urls import reverse_lazy
from django.urls import reverse
from django.contrib.auth.models import User, Group
from django.db import IntegrityError
from django.db.models import Count, Sum, Avg, Min, Max
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.paginator import Paginator
from .models import ResultCategory, TaskCategory, \
    Item, Grade, Task, Inspection, PicRecord
from .forms import ResultCategoryForm, TaskCategoryForm, \
    ItemForm, GradeForm, TaskForm, InspectionForm, PicRecordForm, \
    ImageForm

import datetime, os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date,timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import queue, threading
from collections import defaultdict

#画像系モジュール
from ontheedgeproject.settings import MEDIA_ROOT


global chkbreak

#■■■■■■■■■■■　アカウント認証　■■■■■■■■■■■
def signupview(request): #サインアップ
    if request.method == 'POST':
        email_data = request.POST['email_data']
        password_data = request.POST['password_data']
        try:
            user = User.objects.create_user(email_data, email_data, password_data)
            return redirect('login')
        except IntegrityError:
            return render(request, 'signup.html', {'error':'このユーザーは既に登録されております。'})
    else:
        print(User.objects.all())
        #return render(request, 'signup.html', {})
    return render(request, 'signup.html', {})

def loginview(request): #ログイン画面
    if request.method == 'POST':
        username_data = request.POST['username_data']
        password_data = request.POST['password_data']
        user = authenticate(request, username=username_data, password=password_data)
        if user is not None:
            login(request, user)
            return redirect('headline')
        else:
            return redirect('login')
    return render(request, 'login.html')

def logoutview(request):
    logout(request)
    return redirect('login')

def headlineview(request): #ヘッドラインを表示する
    global chkbreak
    chkbreak=True
    return render(request, 'headline.html')

def datalistview(request): #ヘッドラインを表示する
    return render(request, 'datalist.html')

def setlistview(request): #ヘッドラインを表示する
    return render(request, 'setlist.html')


#■■■■■■■■■■■ 検査結果カテゴリー ■■■■■■■■■■■
def resultcategorylistview(request):
    object = ResultCategory.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'resultcategorylist.html', params)

def createresultcategoryview(request):
    if request.method == 'POST':
        object = ResultCategory() 
        resultcategory = ResultCategoryForm(request.POST, instance=object)
        resultcategory.save()
        return redirect(to='resultcategorylist')
    params = {
        'form': ResultCategoryForm(),
    } 
    return render(request, 'createresultcategory.html', params)

def updateresultcategoryview(request, pk):
    object = ResultCategory.objects.get(pk=pk)
    if request.method == 'POST':
        resultcategory = ResultCategoryForm(request.POST, instance=object)
        resultcategory.save()
        return redirect(to='resultcategorylist')
    params = {
        'pk':pk,
        'object':object,
        'form': ResultCategoryForm(instance=object)
    }
    return render(request, 'updateresultcategory.html', params)

class deleteresultcategory(DeleteView):
    template_name = 'delete.html'
    model = ResultCategory
    success_url = reverse_lazy('resultcategorylist')

#■■■■■■■■■■■ タスクカテゴリー ■■■■■■■■■■■
def taskcategorylistview(request):
    object = TaskCategory.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'taskcategorylist.html', params)

def createtaskcategoryview(request):
    if request.method == 'POST':
        object = TaskCategory() 
        taskcategory = TaskCategoryForm(request.POST, instance=object)
        taskcategory.save()
        return redirect(to='taskcategorylist')
    params = {
        'form': TaskCategoryForm(),
    } 
    return render(request, 'createtaskcategory.html', params)

def updatetaskcategoryview(request, pk):
    object = TaskCategory.objects.get(pk=pk)
    if request.method == 'POST':
        taskcategory = TaskCategoryForm(request.POST, instance=object)
        taskcategory.save()
        return redirect(to='taskcategorylist')
    params = {
        'pk':pk,
        'object':object,
        'form': TaskCategoryForm(instance=object)
    }
    return render(request, 'updatetaskcategory.html', params)

class deletetaskcategory(DeleteView):
    template_name = 'delete.html'
    model = TaskCategory
    success_url = reverse_lazy('taskcategorylist')

#■■■■■■■■■■■ 商品 ■■■■■■■■■■■
def itemlistview(request):
    object = Item.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'itemlist.html', params)

def createitemview(request):
    if request.method == 'POST':
        object = Item() 
        item = ItemForm(request.POST, instance=object)
        item.save()
        return redirect(to='itemlist')
    params = {
        'form': ItemForm(),
    } 
    return render(request, 'createitem.html', params)

def updateitemview(request, pk):
    object = Item.objects.get(pk=pk)
    if request.method == 'POST':
        item = ItemForm(request.POST, instance=object)
        item.save()
        return redirect(to='itemlist')
    params = {
        'pk':pk,
        'object':object,
        'form': ItemForm(instance=object)
    }
    return render(request, 'updateitem.html', params)

class deleteitem(DeleteView):
    template_name = 'delete.html'
    model = Item
    success_url = reverse_lazy('itemlist')

#■■■■■■■■■■■ 商品グレード ■■■■■■■■■■■
def gradelistview(request):
    object = Grade.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'gradelist.html', params)

def creategradeview(request):
    if request.method == 'POST':
        object = Grade() 
        grade = GradeForm(request.POST, instance=object)
        grade.save()
        return redirect(to='gradelist')
    params = {
        'form': GradeForm(),
    } 
    return render(request, 'creategrade.html', params)

def updategradeview(request, pk):
    object = Grade.objects.get(pk=pk)
    if request.method == 'POST':
        grade = GradeForm(request.POST, instance=object)
        grade.save()
        return redirect(to='gradelist')
    params = {
        'pk':pk,
        'object':object,
        'form': GradeForm(instance=object)
    }
    return render(request, 'updategrade.html', params)

class deletegrade(DeleteView):
    template_name = 'delete.html'
    model = Grade
    success_url = reverse_lazy('gradelist')


#■■■■■■■■■■■ タスク ■■■■■■■■■■■
def tasklistview(request):
    object = Task.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'tasklist.html', params)

def createtaskview(request):
    if request.method == 'POST':
        object = Task() 
        task = TaskForm(request.POST, instance=object)
        task.save()
        return redirect(to='tasklist')
    params = {
        'form': TaskForm(),
    } 
    return render(request, 'createtask.html', params)

def updatetaskview(request, pk):
    object = Task.objects.get(pk=pk)
    if request.method == 'POST':
        task = TaskForm(request.POST, instance=object)
        task.save()
        return redirect(to='tasklist')
    inspection = Inspection.objects.filter(task=pk)
    params = {
        'pk':pk,
        'object':object,
        'form': TaskForm(instance=object),
        'inspection':inspection,
    }
    return render(request, 'updatetask.html', params)

class deletetask(DeleteView):
    template_name = 'delete.html'
    model = Task
    success_url = reverse_lazy('tasklist')

#■■■■■■■■■■■ 検査 ■■■■■■■■■■■
def inspectionlistview(request):
    object = Inspection.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'Inspectionlist.html', params)

def back2taskview(request, pk):
    inspection = Inspection.objects.get(pk=pk)
    task = Task.objects.get(pk=inspection.task.id)
    return redirect(to='updatetask', pk=task.id) 

def updateinspectionview(request, pk):
    object = Inspection.objects.get(pk=pk)
    inppath = object.picpath
    #trainpt = analyze(inppath)
    #print(trainpt)
    if request.method == 'POST':
        inspection = InspectionForm(request.POST, instance=object)
        inspection.save()
        print('検査結果登録')
        task = Task.objects.get(pk=object.task.id)
        return redirect(to='updateinspection', pk=pk)
    picrecord = PicRecord.objects.filter(inspection=object.id)
    params = {
        'pk':pk,
        'object':object,
        'form': InspectionForm(instance=object),
        'picrecord':picrecord,
    }
    return render(request, 'updateinspection.html', params)

def failreportview(request):
    #---------- 出力データ生成 ----------
    faildict = defaultdict(list)
    dt_now = datetime.datetime.now()
    today = datetime.date(int(dt_now.year), int(dt_now.month), int(dt_now.day))
    one_month_ago = today - relativedelta(months=1)
    stdy = one_month_ago.replace(day=1)
    task = Task.objects.filter(taskdate__range=[stdy, today])
    label = '令和' + str(stdy.year-2018) + '年'
    label += str(stdy.month) + '-'
    label += str(today.month) + '月度'
    for t in task:
        item = Item.objects.get(pk=t.item.id)
        faildict[t.id].append(item.itemname) #0:商品名
        grade = Grade.objects.get(pk=t.grade.id)
        faildict[t.id].append(grade.gradename) #1:グレード
        faildict[t.id].append(t.itemsize) #2:サイズ
        faildict[t.id].append(t.lot) #3:製造番号
        faildict[t.id].append(t.impdate) #4:入荷日
        faildict[t.id].append(t.taskdate) #5:処理日
        taskcategory = TaskCategory.objects.get(pk=t.taskcategory.id)
        faildict[t.id].append(taskcategory.taskcategory) #6:処理カテゴリー
        faildict[t.id].append(t.tasknum) #7:処理番号
        inspection = Inspection.objects.filter(task=t)
        faildict[t.id].append(len(inspection)) #8:処理枚数
        #不良枚数演算
        shiro_qt=0
        kuro_qt=0
        hoka_qt=0
        gass_qt=0
        #9: 白斑点
        chkid = ResultCategory.objects.get(resultcategory='白斑点')
        shiro = inspection.filter(resultcategory=chkid)
        faildict[t.id].append(len(shiro))

        #10: 黒斑点
        chkid = ResultCategory.objects.get(resultcategory='黒斑点')
        kuro = inspection.filter(resultcategory=chkid)
        faildict[t.id].append(len(kuro))

        #11: 他
        chkid = ResultCategory.objects.get(resultcategory='他不良')
        hoka = inspection.filter(resultcategory=chkid)
        faildict[t.id].append(len(hoka))

        #12: ガスムラ
        chkid = ResultCategory.objects.get(resultcategory='ガスムラ')
        gass = inspection.filter(resultcategory=chkid)
        faildict[t.id].append(len(gass))
    # ---------- 出力 ----------
    book = openpyxl.load_workbook(os.path.join(MEDIA_ROOT, 'failreport.xlsx'))
    sheet = book.worksheets[0]
    sheet['B1'] = label
    line = 4
    for key in faildict.keys():
        sheet['B' + str(line+1)] = '■'
        if faildict[key][0] == '北京鍋': sheet['D' + str(line)] = '■'
        if faildict[key][0] == '炒め鍋': sheet['D' + str(line+1)] = '■'
        if faildict[key][0] == 'マイティ': sheet['D' + str(line+2)] = '■'
        if faildict[key][0] == 'オムレツ': sheet['F' + str(line)] = '■'
        if faildict[key][0] == 'エッグ': sheet['F' + str(line+1)] = '■'
        if faildict[key][0] == '他': sheet['F' + str(line+2)] = '■'

        if faildict[key][1] == '元祖': sheet['H' + str(line)] = '■'
        if faildict[key][1] == 'クックパル': sheet['H' + str(line+1)] = '■'
        if faildict[key][1] == '他': sheet['H' + str(line+2)] = '■'
        if faildict[key][1] == 'プレミアム': sheet['J' + str(line)] = '■'
        if faildict[key][1] == 'ワンコ': sheet['J' + str(line+1)] = '■'

        sheet['L' + str(line + 1)] = faildict[key][2] #サイズ
        sheet['N' + str(line + 1)] = faildict[key][3] #製造番号
        sheet['O' + str(line + 1)] = str(faildict[key][4].month) + '/' + str(faildict[key][4].day) #入荷日
        sheet['P' + str(line + 1)] = str(faildict[key][5].month) + '/' + str(faildict[key][5].day) #入荷日
        sheet['Q' + str(line)] = faildict[key][6] #処理カテゴリー
        sheet['Q' + str(line+1)] = faildict[key][7] #処理番号
        sheet['R' + str(line+1)] = faildict[key][8] #処理枚数
        sheet['T' + str(line+1)] = faildict[key][9] #白斑点数
        sheet['U' + str(line+1)] = faildict[key][10] # 黒斑点数
        sheet['V' + str(line+1)] = faildict[key][11] #他不良数
        sheet['W' + str(line+1)] = faildict[key][12] #他不良数
        line += 3
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=%s" % "異常報告書.xlsx"
    book.save(response)

    return response

    # ---------- ページの表示 ----------
    object = Task.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'tasklist.html', params)

class deleteinspection(DeleteView):
    template_name = 'delete.html'
    model = Inspection
    success_url = reverse_lazy('inspectionlist')

#■■■■■■■■■■■ 画像記録 ■■■■■■■■■■■
def picrecordlistview(request):
    object = PicRecord.objects.all()
    params = {
        'object':object,
    }
    return render(request, 'picrecordlist.html', params)

def updatepicrecordview(request, pk):
    object = picrecord.objects.get(pk=pk)
    if request.method == 'POST':
        picrecord = PicRecordForm(request.POST, instance=object)
        picrecord.save()
        return redirect(to='picrecordlist')
    params = {
        'pk':pk,
        'object':object,
        'form': PicRecordForm(instance=object)
    }
    return render(request, 'updatepicrecord.html', params)

class deletepicrecord(DeleteView):
    template_name = 'delete.html'
    model = PicRecord
    success_url = reverse_lazy('picrecordlist')

